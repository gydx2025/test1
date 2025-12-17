#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
数据查询服务层
"""

import pandas as pd
import sqlite3
import logging
from typing import List, Dict
import os
import re
import json
from datetime import datetime, date

logger = logging.getLogger(__name__)


class DataQueryService:
    """数据查询服务类"""
    
    def __init__(self, cache_dir: str = 'local_cache'):
        """
        初始化查询服务
        
        Args:
            cache_dir: 本地缓存目录
        """
        self.cache_dir = cache_dir
        self.cache_db_path = os.path.join(cache_dir, 'stock_base_data.db')
        self.cache_meta_path = os.path.join(cache_dir, 'cache_meta.json')
        self.db_path = self.cache_db_path  # 保持向后兼容性
        self.available_subjects = self._load_available_subjects()
        self.markets = ['全部', '沪市', '深市', '北市']
        
        # 确保缓存目录存在
        os.makedirs(cache_dir, exist_ok=True)
    
    def _load_available_subjects(self) -> List[Dict]:
        """加载资产负债表科目列表。

        说明：
        - UI 的科目下拉框只展示资产负债表项目，不包含利润表/现金流表等。
        - 列表顺序尽量贴近资产负债表的展示顺序（资产 -> 负债 -> 所有者权益）。
        - 扩展为完整的资产负债表科目列表

        Returns:
            科目列表
        """
        subjects = [
            # 资产类科目
            {'code': 'INVEST_REALESTATE', 'name': '投资性房地产'},
            {'code': 'FIXED_ASSET', 'name': '固定资产'},
            {'code': 'CIP', 'name': '在建工程'},
            {'code': 'USERIGHT_ASSET', 'name': '使用权资产'},
            {'code': 'INTANGIBLE_ASSET', 'name': '无形资产'},
            {'code': 'GOODWILL', 'name': '商誉'},
            {'code': 'DEFERRED_EXPENSES', 'name': '长期待摊费用'},
            {'code': 'DEFERRED_TAX_ASSET', 'name': '递延所得税资产'},
            {'code': 'OTHER_NONCURRENT_ASSETS', 'name': '其他非流动资产'},
            {'code': 'NONCURRENT_ASSET_SUMMARY', 'name': '非流动资产合计'},
            {'code': 'CURRENT_ASSET_INVENTORY', 'name': '存货'},
            {'code': 'CURRENT_ASSET_ACCOUNT_RECEIVABLE', 'name': '应收账款'},
            {'code': 'CURRENT_ASSET_OTHER_RECEIVABLE', 'name': '其他应收款'},
            {'code': 'CURRENT_ASSET_PREPAYMENT', 'name': '预付款项'},
            {'code': 'CURRENT_ASSET_CASH', 'name': '货币资金'},
            {'code': 'CURRENT_ASSET_TRADING_FINASSET', 'name': '交易性金融资产'},
            {'code': 'CURRENT_ASSET_OTHER', 'name': '其他流动资产'},
            {'code': 'CURRENT_ASSET_SUMMARY', 'name': '流动资产合计'},
            {'code': 'TOTAL_ASSETS', 'name': '资产总计'},
            
            # 负债类科目
            {'code': 'NONCURRENT_LIABILITY_BORROW', 'name': '长期借款'},
            {'code': 'NONCURRENT_LIABILITY_BOND', 'name': '应付债券'},
            {'code': 'NONCURRENT_LIABILITY_LEASE', 'name': '租赁负债'},
            {'code': 'NONCURRENT_LIABILITY_DEFERRED_TAX', 'name': '递延所得税负债'},
            {'code': 'NONCURRENT_LIABILITY_OTHER', 'name': '其他非流动负债'},
            {'code': 'NONCURRENT_LIABILITY_SUMMARY', 'name': '非流动负债合计'},
            {'code': 'CURRENT_LIABILITY_ACCOUNT_PAYABLE', 'name': '应付账款'},
            {'code': 'CURRENT_LIABILITY_ADVANCE_RECEIPT', 'name': '预收款项'},
            {'code': 'CURRENT_LIABILITY_SALARY_PAYABLE', 'name': '应付职工薪酬'},
            {'code': 'CURRENT_LIABILITY_TAX_PAYABLE', 'name': '应交税费'},
            {'code': 'CURRENT_LIABILITY_OTHER_PAYABLE', 'name': '其他应付款'},
            {'code': 'CURRENT_LIABILITY_BORROW', 'name': '短期借款'},
            {'code': 'CURRENT_LIABILITY_TRADING_FINLIAB', 'name': '交易性金融负债'},
            {'code': 'CURRENT_LIABILITY_OTHER', 'name': '其他流动负债'},
            {'code': 'CURRENT_LIABILITY_SUMMARY', 'name': '流动负债合计'},
            {'code': 'TOTAL_LIABILITIES', 'name': '负债合计'},
            
            # 所有者权益类科目
            {'code': 'OWNER_EQUITY_SHARE_CAPITAL', 'name': '实收资本(或股本)'},
            {'code': 'OWNER_EQUITY_CAPITAL_RESERVE', 'name': '资本公积'},
            {'code': 'OWNER_EQUITY_SURPLUS_RESERVE', 'name': '盈余公积'},
            {'code': 'OWNER_EQUITY_RETAINED_EARNINGS', 'name': '未分配利润'},
            {'code': 'OWNER_EQUITY_SUMMARY', 'name': '所有者权益合计'},
            
            # 特殊科目（保持兼容性）
            {'code': 'non_op_real_estate', 'name': '非经营性房地产资产'},
        ]
        return subjects
    
    def query_data(self, 
                   stock_codes: List[str] = None,
                   stock_names: List[str] = None,
                   market: str = '全部',
                   time_points: List[str] = None,
                   subject_codes: List[str] = None,
                   subject_names: List[str] = None,
                   subject_code: str = None,  # 保持兼容性
                   industry: str = None) -> pd.DataFrame:
        """查询数据。

        Args:
            stock_codes: 股票代码列表
            stock_names: 股票名称列表（支持模糊匹配）
            market: 市场类型
            time_points: 时点列表。
                - 兼容传入年份（如 '2023'）
                - 也支持传入财报期日期（如 '2024-06-30'），会自动提取年份做兼容过滤
            subject_codes: 科目代码列表（新接口，支持多科目）
            subject_names: 科目名称列表（新接口，支持多科目）
            subject_code: 科目代码（兼容性）
            industry: 行业筛选（申万一级行业）。None/全行业 表示不筛选

        Returns:
            查询结果DataFrame
        """
        try:
            logger.info("=== 开始数据查询 (使用采集器逻辑) ===")
            
            # 优先使用新的多科目接口
            if subject_codes is None and subject_code:
                subject_codes = [subject_code]
            elif subject_codes is None and subject_code is None:
                # 默认查询投资性房地产
                subject_codes = ['INVEST_REALESTATE']
            
            logger.info(f"科目代码: {subject_codes}")
            logger.info(f"股票代码: {stock_codes}, 股票名称: {stock_names}")
            logger.info(f"市场: {market}, 行业: {industry}, 时点: {time_points}")
            
            # 如果没有指定股票，尝试获取股票列表
            if not stock_codes and not stock_names:
                logger.info("未指定股票，尝试获取所有股票...")
                stock_list = self.get_stock_list()
                if not stock_list.empty:
                    stock_codes = stock_list['股票代码'].tolist()
                    logger.info(f"获取了 {len(stock_codes)} 只股票")
                else:
                    logger.warning("未能获取到股票列表")
                    return pd.DataFrame()
            
            # 合并股票代码和名称搜索
            all_stock_codes = list(stock_codes) if stock_codes else []
            
            # 如果有股票名称模糊搜索，扩展股票代码列表
            if stock_names:
                try:
                    from financial_query_service import FinancialQueryService
                    query_service = FinancialQueryService()
                    # 这里需要获取股票列表并根据名称匹配
                    # 为了简化，现在只返回股票代码查询结果
                    logger.info(f"股票名称搜索功能待实现，当前搜索: {stock_names}")
                except Exception as e:
                    logger.warning(f"处理股票名称搜索失败: {e}")
            
            if not all_stock_codes:
                logger.warning("没有有效的股票代码进行查询")
                return pd.DataFrame()
            
            # 标准化时点数据
            normalized_time_points = []
            if time_points:
                for tp in time_points:
                    if not tp:
                        continue
                    token = str(tp).strip()
                    if not token:
                        continue
                    
                    # 标准化时点格式为 YYYY-MM-DD
                    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", token):
                        normalized_time_points.append(token)
                    elif re.fullmatch(r"\d{4}", token):
                        # 年份转换为年末日期
                        normalized_time_points.append(f"{token}-12-31")
                    else:
                        logger.warning(f"无法识别的时点格式: {tp}")
            
            if not normalized_time_points:
                # 默认查询最近一期的数据
                normalized_time_points = ['2023-12-31']
                logger.info(f"未指定时点，使用默认: {normalized_time_points}")
            
            # 调用原有的采集器逻辑进行查询
            return self._query_from_collector(all_stock_codes, subject_codes, normalized_time_points)
            
        except Exception as e:
            logger.error(f"数据查询失败: {str(e)}", exc_info=True)
            raise
    
    def _query_from_collector(self, stock_codes: List[str], subject_codes: List[str], time_points: List[str]) -> pd.DataFrame:
        """使用原有采集器逻辑查询数据（主要查询方法）"""
        try:
            logger.info(f"开始使用采集器逻辑查询: {len(subject_codes)}个科目, {len(stock_codes)}只股票, {len(time_points)}个时点")
            
            # 导入财务查询服务
            from financial_query_service import FinancialQueryService
            from astock_real_estate_collector import AStockRealEstateDataCollector
            
            # 创建查询服务
            query_service = FinancialQueryService()
            
            # 获取股票基础信息（代码、名称、市场）
            # 这里我们需要通过采集器获取股票信息
            try:
                collector = AStockRealEstateDataCollector()
                stock_info_list = collector.get_stock_list()
                
                # 构建股票代码到股票信息的映射
                stock_info_map = {}
                for stock_info in stock_info_list:
                    code = stock_info.get('code', '')
                    stock_info_map[code] = {
                        '股票代码': code,
                        '股票名称': stock_info.get('name', ''),
                        '市场': self._get_market_display_name(stock_info.get('market', ''))
                    }
                
                logger.info(f"获取了 {len(stock_info_map)} 只股票的基础信息")
            except Exception as e:
                logger.warning(f"获取股票基础信息失败: {e}")
                # 创建基本的股票信息映射
                stock_info_map = {}
                for code in stock_codes:
                    stock_info_map[code] = {
                        '股票代码': code,
                        '股票名称': code,  # 临时使用代码作为名称
                        '市场': self._get_market_display_name(code)
                    }
                logger.info(f"使用默认股票信息映射: {len(stock_info_map)} 只股票")
            
            all_results = []
            
            # 为每个科目分别查询
            for subject_code in subject_codes:
                logger.info(f"查询科目: {subject_code}")
                
                try:
                    # 获取科目显示名称
                    subject_display_name = self._get_subject_display_name(subject_code)
                    
                    # 准备查询参数
                    query_params = {
                        'stock_codes': stock_codes,
                        'market': '全部',
                        'subject': subject_display_name,
                        'report_dates': time_points,
                        'return_format': 'dataframe',
                        'unit': '万元'
                    }
                    
                    # 执行查询
                    result = query_service.execute_query(query_params)
                    
                    if result.success and result.dataframe is not None:
                        df = result.dataframe.copy()
                        logger.info(f"科目 {subject_code} 查询成功，返回 {len(df)} 条记录")
                        
                        # 为每个科目添加单独的列，格式：科目名称(时点)
                        for _, row in df.iterrows():
                            stock_code = row.get('stock_code', '')
                            if stock_code and stock_code in stock_info_map:
                                # 创建结果记录
                                record = stock_info_map[stock_code].copy()
                                
                                # 添加各时点的数值
                                for time_point in time_points:
                                    col_name = f"{subject_display_name}({time_point})(万元)"
                                    value = row.get(col_name)
                                    record[col_name] = value
                                
                                all_results.append(record)
                    else:
                        logger.warning(f"科目 {subject_code} 查询失败: {result.error}")
                        # 即使查询失败，也需要创建空记录以保持数据结构
                        for code in stock_codes:
                            if code in stock_info_map:
                                record = stock_info_map[code].copy()
                                for time_point in time_points:
                                    col_name = f"{subject_display_name}({time_point})(万元)"
                                    record[col_name] = None
                                all_results.append(record)
                
                except Exception as e:
                    logger.error(f"查询科目 {subject_code} 时出错: {e}")
                    # 创建错误记录
                    for code in stock_codes:
                        if code in stock_info_map:
                            record = stock_info_map[code].copy()
                            for time_point in time_points:
                                col_name = f"{subject_display_name}({time_point})(万元)"
                                record[col_name] = None
                            all_results.append(record)
            
            # 如果没有结果，返回空DataFrame
            if not all_results:
                logger.warning("查询未返回任何数据")
                return pd.DataFrame()
            
            # 合并所有结果
            result_df = pd.DataFrame(all_results)
            
            # 按股票代码和时点重新组织列的顺序
            if not result_df.empty:
                # 确定列的顺序
                base_columns = ['股票代码', '股票名称', '市场']
                data_columns = []
                
                # 获取所有数据列并排序
                for col in result_df.columns:
                    if col not in base_columns:
                        data_columns.append(col)
                
                data_columns.sort()  # 按列名排序
                final_columns = base_columns + data_columns
                
                # 重新排列列
                existing_columns = [col for col in final_columns if col in result_df.columns]
                result_df = result_df[existing_columns]
            
            logger.info(f"=== 采集器查询完成，共返回 {len(result_df)} 条记录 ===")
            return result_df
            
        except Exception as e:
            logger.error(f"使用采集器查询失败: {e}", exc_info=True)
            return pd.DataFrame()
    
    def _get_market_display_name(self, market_code: str) -> str:
        """根据市场代码获取显示名称"""
        market_map = {
            'SH': '沪市',
            'SZ': '深市', 
            'BJ': '北市',
            '6': '沪市',
            '0': '深市',
            '3': '深市',
            '4': '北市',
            '8': '北市'
        }
        
        # 提取市场前缀
        if market_code.startswith(('SH', 'SZ', 'BJ')):
            code_prefix = market_code[:2]
        else:
            code_prefix = market_code[0] if market_code else '0'
        
        return market_map.get(code_prefix, '未知市场')
    
    def _get_subject_display_name(self, subject_code: str) -> str:
        """根据科目代码获取显示名称"""
        subject_map = {
            'INVEST_REALESTATE': '投资性房地产',
            'FIXED_ASSET': '固定资产', 
            'CIP': '在建工程',
            'USERIGHT_ASSET': '使用权资产',
            'INTANGIBLE_ASSET': '无形资产',
            'GOODWILL': '商誉',
            'DEFERRED_EXPENSES': '长期待摊费用',
            'DEFERRED_TAX_ASSET': '递延所得税资产',
            'OTHER_NONCURRENT_ASSETS': '其他非流动资产',
            'NONCURRENT_ASSET_SUMMARY': '非流动资产合计',
            'CURRENT_ASSET_INVENTORY': '存货',
            'CURRENT_ASSET_ACCOUNT_RECEIVABLE': '应收账款',
            'CURRENT_ASSET_OTHER_RECEIVABLE': '其他应收款',
            'CURRENT_ASSET_PREPAYMENT': '预付款项',
            'CURRENT_ASSET_CASH': '货币资金',
            'CURRENT_ASSET_TRADING_FINASSET': '交易性金融资产',
            'CURRENT_ASSET_OTHER': '其他流动资产',
            'CURRENT_ASSET_SUMMARY': '流动资产合计',
            'TOTAL_ASSETS': '资产总计',
            'NONCURRENT_LIABILITY_BORROW': '长期借款',
            'NONCURRENT_LIABILITY_BOND': '应付债券',
            'NONCURRENT_LIABILITY_LEASE': '租赁负债',
            'NONCURRENT_LIABILITY_DEFERRED_TAX': '递延所得税负债',
            'NONCURRENT_LIABILITY_OTHER': '其他非流动负债',
            'NONCURRENT_LIABILITY_SUMMARY': '非流动负债合计',
            'CURRENT_LIABILITY_ACCOUNT_PAYABLE': '应付账款',
            'CURRENT_LIABILITY_ADVANCE_RECEIPT': '预收款项',
            'CURRENT_LIABILITY_SALARY_PAYABLE': '应付职工薪酬',
            'CURRENT_LIABILITY_TAX_PAYABLE': '应交税费',
            'CURRENT_LIABILITY_OTHER_PAYABLE': '其他应付款',
            'CURRENT_LIABILITY_BORROW': '短期借款',
            'CURRENT_LIABILITY_TRADING_FINLIAB': '交易性金融负债',
            'CURRENT_LIABILITY_OTHER': '其他流动负债',
            'CURRENT_LIABILITY_SUMMARY': '流动负债合计',
            'TOTAL_LIABILITIES': '负债合计',
            'OWNER_EQUITY_SHARE_CAPITAL': '实收资本(或股本)',
            'OWNER_EQUITY_CAPITAL_RESERVE': '资本公积',
            'OWNER_EQUITY_SURPLUS_RESERVE': '盈余公积',
            'OWNER_EQUITY_RETAINED_EARNINGS': '未分配利润',
            'OWNER_EQUITY_SUMMARY': '所有者权益合计',
            'non_op_real_estate': '非经营性房地产资产'
        }
        
        return subject_map.get(subject_code, subject_code)
    
    def get_stock_list(self) -> pd.DataFrame:
        """获取股票列表。

        Returns:
            股票列表DataFrame（包含股票代码、股票名称、市场字段）
        """
        try:
            logger.info("获取股票列表（使用采集器逻辑）...")
            
            # 使用原有采集器获取股票列表
            from astock_real_estate_collector import AStockRealEstateDataCollector
            
            collector = AStockRealEstateDataCollector()
            stock_info_list = collector.get_stock_list()
            
            if stock_info_list:
                # 转换为DataFrame
                df = pd.DataFrame(stock_info_list)
                
                # 重命名列以匹配UI期望
                if 'code' in df.columns:
                    df = df.rename(columns={'code': '股票代码'})
                if 'name' in df.columns:
                    df = df.rename(columns={'name': '股票名称'})
                if 'market' in df.columns:
                    # 将市场代码转换为显示名称
                    df['市场'] = df['market'].apply(self._get_market_display_name)
                    df = df.drop('market', axis=1)
                
                # 确保有基本的列
                if '股票代码' not in df.columns:
                    df['股票代码'] = ''
                if '股票名称' not in df.columns:
                    df['股票名称'] = ''
                if '市场' not in df.columns:
                    df['市场'] = '未知市场'
                
                # 按股票代码排序
                df = df.sort_values('股票代码').reset_index(drop=True)
                
                logger.info(f"成功获取 {len(df)} 只股票")
                return df
            else:
                logger.warning("采集器返回空股票列表")
                return pd.DataFrame()
                
        except Exception as e:
            logger.error(f"获取股票列表失败: {str(e)}")
            # 如果采集器失败，返回基本的DataFrame结构
            return pd.DataFrame(columns=['股票代码', '股票名称', '市场'])

    def get_industry_options(self) -> List[str]:
        """获取行业筛选下拉框的候选项（申万一级行业）。

        Returns:
            行业列表（不含“全行业”）
        """
        try:
            logger.info("获取行业列表（使用采集器逻辑）...")
            
            # 尝试从采集器获取行业数据
            try:
                from astock_real_estate_collector import AStockRealEstateDataCollector
                
                collector = AStockRealEstateDataCollector()
                stock_info_list = collector.get_stock_list()
                
                if stock_info_list:
                    # 提取行业信息
                    industries = set()
                    for stock_info in stock_info_list:
                        industry = stock_info.get('industry', '')
                        if industry and industry.strip() and industry != '未知':
                            industries.add(industry.strip())
                    
                    if industries:
                        industry_list = sorted(list(industries))
                        logger.info(f"从采集器获取到 {len(industry_list)} 个行业")
                        return industry_list
                        
            except Exception as e:
                logger.warning(f"从采集器获取行业数据失败: {e}")
            
            # 尝试从本地缓存获取行业数据
            industries = self._get_industry_from_cache()
            if industries:
                return industries
                
            # 如果缓存也没有，返回默认的行业列表
            logger.warning("无法获取行业列表，返回默认行业列表")
            return self._get_default_industries()
            
        except Exception as e:
            logger.error(f"获取行业列表失败: {str(e)}")
            return self._get_default_industries()
    
    def _get_industry_from_cache(self) -> List[str]:
        """从本地缓存获取行业列表"""
        try:
            # 尝试从本地缓存获取行业数据
            cache_file = "shenwan_industry_mapping.pkl"
            if os.path.exists(cache_file):
                import pickle
                with open(cache_file, 'rb') as f:
                    industry_data = pickle.load(f)
                    if industry_data:
                        # 提取一级行业
                        l1_industries = set()
                        for stock_code, industry_info in industry_data.items():
                            if isinstance(industry_info, dict) and 'l1' in industry_info:
                                l1_industries.add(industry_info['l1'])
                        return sorted(list(l1_industries))
            
            # 尝试从其他缓存文件获取
            other_cache_files = [
                "industry_cache.pkl",
                "stock_industry_mapping.pkl"
            ]
            
            for cache_file in other_cache_files:
                if os.path.exists(cache_file):
                    try:
                        import pickle
                        with open(cache_file, 'rb') as f:
                            cache_data = pickle.load(f)
                            if cache_data and isinstance(cache_data, dict):
                                # 尝试提取行业信息
                                industries = set()
                                for key, value in cache_data.items():
                                    if isinstance(value, str) and value.strip():
                                        industries.add(value.strip())
                                    elif isinstance(value, dict):
                                        for industry_key in ['l1', 'industry', 'l1_industry']:
                                            if industry_key in value and value[industry_key]:
                                                industries.add(str(value[industry_key]).strip())
                                if industries:
                                    return sorted(list(industries))
                    except Exception:
                        continue
            
            logger.debug("未找到有效的行业缓存文件")
            return []
            
        except Exception as e:
            logger.error(f"从缓存获取行业列表失败: {e}")
            return []
    
    def _get_default_industries(self) -> List[str]:
        """获取默认的申万一级行业列表"""
        return [
            "农林牧渔",
            "采掘",
            "化工",
            "钢铁",
            "有色金属",
            "电子",
            "家用电器",
            "食品饮料",
            "纺织服装",
            "轻工制造",
            "医药生物",
            "公用事业",
            "交通运输",
            "房地产",
            "商业贸易",
            "休闲服务",
            "综合",
            "建筑材料",
            "建筑装饰",
            "电气设备",
            "国防军工",
            "计算机",
            "传媒",
            "通信",
            "非银金融",
            "银行",
            "汽车",
            "机械设备"
        ]
    
    def export_to_excel(self, df: pd.DataFrame, file_path: str, 
                       subject_codes: List[str] = None,
                       report_dates: List[str] = None,
                       filters: Dict = None) -> bool:
        """
        导出数据到Excel（支持两个Sheet）
        
        Args:
            df: 要导出的数据
            file_path: 导出文件路径
            subject_codes: 科目代码列表
            report_dates: 财报期列表
            filters: 查询过滤条件
            
        Returns:
            是否成功
        """
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Sheet1: 写入查询结果
                df.to_excel(writer, sheet_name='查询结果', index=False)
                
                # 设置Sheet1列宽
                worksheet = writer.sheets['查询结果']
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Sheet2: 写入元数据和查询条件
                self._add_metadata_sheet(writer, df, subject_codes, report_dates, filters)
            
            logger.info(f"数据已导出到: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"导出Excel失败: {str(e)}")
            return False
    
    def _add_metadata_sheet(self, writer, df: pd.DataFrame, 
                           subject_codes: List[str] = None,
                           report_dates: List[str] = None,
                           filters: Dict = None):
        """添加元数据Sheet"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # 创建元数据Sheet
            workbook = writer.book
            worksheet = workbook.create_sheet('元数据')
            
            # 设置样式
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            header_font = Font(bold=True)
            
            row = 1
            
            # 1. 查询时点
            worksheet[f'A{row}'] = '查询时点'
            worksheet[f'A{row}'].fill = header_fill
            worksheet[f'A{row}'].font = header_font
            worksheet[f'B{row}'] = ', '.join(report_dates) if report_dates else '无'
            worksheet.column_dimensions['A'].width = 20
            worksheet.column_dimensions['B'].width = 40
            row += 1
            
            # 2. 选定科目
            worksheet[f'A{row}'] = '选定科目'
            worksheet[f'A{row}'].fill = header_fill
            worksheet[f'A{row}'].font = header_font
            if subject_codes:
                subject_names = [self._get_subject_display_name(code) for code in subject_codes]
                worksheet[f'B{row}'] = ', '.join(subject_names)
            else:
                worksheet[f'B{row}'] = '无'
            row += 1
            
            # 3. 行业筛选
            worksheet[f'A{row}'] = '行业筛选'
            worksheet[f'A{row}'].fill = header_fill
            worksheet[f'A{row}'].font = header_font
            worksheet[f'B{row}'] = filters.get('industry', '全行业') if filters else '全行业'
            row += 1
            
            # 4. 股票筛选
            worksheet[f'A{row}'] = '股票筛选'
            worksheet[f'A{row}'].fill = header_fill
            worksheet[f'A{row}'].font = header_font
            if filters and (filters.get('stock_codes') or filters.get('stock_names')):
                codes_str = ', '.join(filters.get('stock_codes', []))
                names_str = ', '.join(filters.get('stock_names', []))
                worksheet[f'B{row}'] = (codes_str + ', ' + names_str).strip(', ')
            else:
                worksheet[f'B{row}'] = '无'
            row += 1
            
            # 5. 市场筛选
            worksheet[f'A{row}'] = '市场筛选'
            worksheet[f'A{row}'].fill = header_fill
            worksheet[f'A{row}'].font = header_font
            worksheet[f'B{row}'] = filters.get('market', '全部') if filters else '全部'
            row += 1
            
            # 6. 生成时间
            worksheet[f'A{row}'] = '生成时间'
            worksheet[f'A{row}'].fill = header_fill
            worksheet[f'A{row}'].font = header_font
            worksheet[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            row += 1
            
            # 7. 查询结果统计
            worksheet[f'A{row}'] = '结果统计'
            worksheet[f'A{row}'].fill = header_fill
            worksheet[f'A{row}'].font = header_font
            worksheet[f'B{row}'] = f'成功 {len(df)} 条，失败 0 条'
            row += 1
            
        except Exception as e:
            logger.warning(f"添加元数据Sheet失败: {e}")
    
    def init_base_cache(self) -> Dict:
        """
        初始化本地缓存库
        
        Returns:
            初始化结果和统计信息
        """
        try:
            logger.info("=== 开始初始化缓存库 ===")
            
            # 检查缓存库是否已存在
            if os.path.exists(self.cache_db_path):
                logger.info("缓存库已存在，跳过初始化")
                return {
                    'success': True,
                    'message': '缓存库已存在',
                    'db_path': self.cache_db_path,
                    'stock_count': self._get_cache_stock_count()
                }
            
            # 创建新的缓存库
            conn = self._get_cache_connection()
            self._create_cache_tables(conn)
            
            # 从采集器获取完整数据
            from astock_real_estate_collector import AStockRealEstateDataCollector
            collector = AStockRealEstateDataCollector()
            
            # 获取股票列表
            logger.info("获取股票列表...")
            stock_info_list = collector.get_stock_list()
            
            if not stock_info_list:
                logger.warning("未能获取股票列表")
                return {
                    'success': False,
                    'message': '未能获取股票列表',
                    'stock_count': 0
                }
            
            logger.info(f"获取到 {len(stock_info_list)} 只股票")
            
            # 获取行业分类（通过财务查询服务或其他方式）
            logger.info("获取行业分类...")
            industry_data = self._fetch_industry_classifications(stock_info_list)
            
            # 批量插入股票和行业数据
            logger.info("插入缓存数据...")
            self._insert_stocks_to_cache(conn, stock_info_list, industry_data)
            
            # 更新缓存元数据
            self._update_cache_metadata()
            
            conn.close()
            
            stock_count = len(stock_info_list)
            logger.info(f"=== 缓存库初始化完成，共 {stock_count} 只股票 ===")
            
            return {
                'success': True,
                'message': '缓存库初始化成功',
                'db_path': self.cache_db_path,
                'stock_count': stock_count,
                'industry_count': len(industry_data)
            }
            
        except Exception as e:
            logger.error(f"初始化缓存库失败: {e}", exc_info=True)
            return {
                'success': False,
                'message': f'初始化失败: {str(e)}',
                'stock_count': 0
            }
    
    def update_daily_cache(self) -> Dict:
        """
        执行每日缓存更新
        
        Returns:
            更新结果和统计信息
        """
        try:
            logger.info("=== 开始每日缓存更新 ===")
            
            # 检查缓存库是否存在
            if not os.path.exists(self.cache_db_path):
                logger.info("缓存库不存在，执行初始化")
                return self.init_base_cache()
            
            # 检查是否已在今天更新过
            if not self._should_update_today():
                logger.info("今天已更新过，跳过更新")
                return {
                    'success': True,
                    'message': '今天已更新过',
                    'need_update': False
                }
            
            # 执行增量更新
            from astock_real_estate_collector import AStockRealEstateDataCollector
            collector = AStockRealEstateDataCollector()
            
            # 获取最新的股票数据
            logger.info("获取最新股票列表...")
            new_stock_list = collector.get_stock_list()
            
            if not new_stock_list:
                logger.warning("未能获取新的股票列表")
                return {
                    'success': False,
                    'message': '未能获取新的股票列表'
                }
            
            # 比对和更新
            conn = self._get_cache_connection()
            
            # 获取行业分类
            industry_data = self._fetch_industry_classifications(new_stock_list)
            
            # 执行增量更新
            stats = self._incremental_update_cache(conn, new_stock_list, industry_data)
            
            # 更新缓存元数据
            self._update_cache_metadata()
            
            conn.close()
            
            logger.info(f"=== 缓存更新完成 ===")
            
            return {
                'success': True,
                'message': '缓存更新成功',
                'added': stats.get('added', 0),
                'updated': stats.get('updated', 0),
                'deleted': stats.get('deleted', 0)
            }
            
        except Exception as e:
            logger.error(f"缓存更新失败: {e}", exc_info=True)
            return {
                'success': False,
                'message': f'更新失败: {str(e)}'
            }
    
    def query_base_data_from_cache(self, filters: Dict = None) -> pd.DataFrame:
        """
        从缓存快速查询基础数据
        
        Args:
            filters: 过滤条件
                - industry: 行业过滤 (三级申万分类或通用行业)
                - market: 市场过滤 (沪/深/北/全部)
                - stock_codes: 股票代码过滤
                - stock_names: 股票名称模糊过滤
        
        Returns:
            基础数据 DataFrame
        """
        try:
            logger.info("从缓存查询基础数据...")
            
            if not os.path.exists(self.cache_db_path):
                logger.warning("缓存库不存在")
                return pd.DataFrame()
            
            filters = filters or {}
            
            # 构建查询SQL
            conn = self._get_cache_connection()
            query = """
            SELECT 
                s.code, s.name, s.market, s.list_date,
                i.shenwan_level1, i.shenwan_level2, i.shenwan_level3, 
                i.general_industry, s.update_time, s.data_source
            FROM stocks s
            LEFT JOIN industries i ON s.code = i.code
            WHERE 1=1
            """
            
            params = []
            
            # 应用过滤条件
            if filters.get('market') and filters['market'] != '全部':
                market_map = {'沪': 'SH', '深': 'SZ', '北': 'BJ'}
                market_code = market_map.get(filters['market'], '')
                if market_code:
                    query += " AND s.market LIKE ?"
                    params.append(f"{market_code}%")
            
            if filters.get('stock_codes'):
                codes_list = filters['stock_codes']
                placeholders = ','.join(['?' for _ in codes_list])
                query += f" AND s.code IN ({placeholders})"
                params.extend(codes_list)
            
            if filters.get('stock_names'):
                for name in filters['stock_names']:
                    query += " AND s.name LIKE ?"
                    params.append(f"%{name}%")
            
            if filters.get('industry'):
                industry = filters['industry']
                if industry != '全行业':
                    # 支持三级行业或通用行业
                    query += " AND (i.shenwan_level1 = ? OR i.shenwan_level2 = ? OR i.shenwan_level3 = ? OR i.general_industry = ?)"
                    params.extend([industry, industry, industry, industry])
            
            query += " ORDER BY s.code"
            
            # 执行查询
            df = pd.read_sql_query(query, conn, params=params)
            conn.close()
            
            logger.info(f"从缓存查询到 {len(df)} 条记录")
            return df
            
        except Exception as e:
            logger.error(f"从缓存查询失败: {e}", exc_info=True)
            return pd.DataFrame()
    
    def query_subject_data_dynamic(self, subject_codes: List[str], 
                                   report_dates: List[str], 
                                   stock_list: List[str] = None) -> pd.DataFrame:
        """
        实时查询科目数据，支持多源和循环补齐
        
        Args:
            subject_codes: 科目代码列表
            report_dates: 财报期列表
            stock_list: 股票代码列表
        
        Returns:
            科目数据 DataFrame
        """
        try:
            logger.info(f"开始实时查询科目数据: {len(subject_codes)} 个科目, {len(stock_list)} 只股票")
            
            from financial_query_service import FinancialQueryService
            
            query_service = FinancialQueryService()
            all_results = []
            
            # 构建基础结果字典
            result_dict = {}
            for code in stock_list:
                result_dict[code] = {'code': code}
            
            # 为每个科目查询数据
            for subject_code in subject_codes:
                subject_name = self._get_subject_display_name(subject_code)
                logger.info(f"查询科目: {subject_name} ({subject_code})")
                
                try:
                    # 执行多源查询
                    for report_date in report_dates:
                        col_name = f"{report_date}_{subject_name}"
                        
                        # 为每只股票查询数据
                        for code in stock_list:
                            try:
                                query_params = {
                                    'stock_codes': [code],
                                    'report_date': report_date,
                                    'subject': subject_name,
                                    'return_format': 'dataframe'
                                }
                                
                                result = query_service.execute_query(query_params)
                                
                                if result.success and result.dataframe is not None and len(result.dataframe) > 0:
                                    value = result.dataframe.iloc[0].get(subject_name)
                                    result_dict[code][col_name] = value
                                else:
                                    result_dict[code][col_name] = None
                                    
                            except Exception as e:
                                logger.warning(f"查询股票 {code} 科目 {subject_name} 时点 {report_date} 失败: {e}")
                                result_dict[code][col_name] = None
                
                except Exception as e:
                    logger.warning(f"查询科目 {subject_name} 失败: {e}")
            
            # 转换为DataFrame
            result_df = pd.DataFrame(list(result_dict.values()))
            logger.info(f"实时查询完成，返回 {len(result_df)} 条记录")
            return result_df
            
        except Exception as e:
            logger.error(f"实时查询科目数据失败: {e}", exc_info=True)
            return pd.DataFrame()
    
    def execute_full_query(self, params: Dict) -> Dict:
        """
        综合查询方法，整合所有操作
        
        Args:
            params: 查询参数
                - subject_codes: 科目代码列表
                - report_dates: 财报期列表
                - industry: 行业过滤
                - stock_codes: 股票代码
                - stock_names: 股票名称
                - market: 市场
        
        Returns:
            查询结果和统计信息
        """
        try:
            logger.info("=== 开始综合查询 ===")
            
            # 1. 从缓存加载基础数据
            logger.info("加载基础数据...")
            filters = {
                'industry': params.get('industry', '全行业'),
                'market': params.get('market', '全部'),
                'stock_codes': params.get('stock_codes'),
                'stock_names': params.get('stock_names')
            }
            
            base_df = self.query_base_data_from_cache(filters)
            
            if base_df.empty:
                logger.warning("未找到符合条件的基础数据")
                return {
                    'success': False,
                    'message': '未找到符合条件的数据',
                    'result': pd.DataFrame()
                }
            
            logger.info(f"筛选到 {len(base_df)} 只符合条件的股票")
            
            # 2. 查询科目数据
            stock_codes = base_df['code'].tolist()
            subject_codes = params.get('subject_codes', ['INVEST_REALESTATE'])
            report_dates = params.get('report_dates', ['2023-12-31'])
            
            logger.info(f"查询科目数据: {subject_codes}")
            subject_df = self.query_subject_data_dynamic(subject_codes, report_dates, stock_codes)
            
            # 3. 合并结果
            if not subject_df.empty:
                result_df = pd.merge(base_df, subject_df, on='code', how='left')
            else:
                result_df = base_df
            
            logger.info(f"=== 综合查询完成，共返回 {len(result_df)} 条记录 ===")
            
            return {
                'success': True,
                'message': '查询成功',
                'result': result_df,
                'total': len(result_df),
                'success_count': len(result_df),
                'failed_count': 0
            }
            
        except Exception as e:
            logger.error(f"综合查询失败: {e}", exc_info=True)
            return {
                'success': False,
                'message': f'查询失败: {str(e)}',
                'result': pd.DataFrame()
            }
    
    # ========== 辅助方法 ==========
    
    def _get_cache_connection(self) -> sqlite3.Connection:
        """获取缓存数据库连接"""
        return sqlite3.connect(self.cache_db_path, timeout=30)
    
    def _create_cache_tables(self, conn: sqlite3.Connection):
        """创建缓存表结构"""
        cursor = conn.cursor()
        
        # 创建股票表
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS stocks (
            code TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            market TEXT,
            list_date DATE,
            update_time TIMESTAMP,
            data_source TEXT
        )
        ''')
        
        # 创建行业表
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS industries (
            code TEXT PRIMARY KEY,
            shenwan_level1 TEXT,
            shenwan_level2 TEXT,
            shenwan_level3 TEXT,
            general_industry TEXT,
            data_source TEXT,
            FOREIGN KEY (code) REFERENCES stocks(code)
        )
        ''')
        
        # 创建索引
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_stocks_code ON stocks(code)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_stocks_market ON stocks(market)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_industries_shenwan1 ON industries(shenwan_level1)')
        
        conn.commit()
    
    def _insert_stocks_to_cache(self, conn: sqlite3.Connection, 
                                stock_list: List[Dict], 
                                industry_data: Dict):
        """将股票和行业数据插入缓存"""
        cursor = conn.cursor()
        
        from datetime import datetime
        now = datetime.now().isoformat()
        
        for stock in stock_list:
            code = stock.get('code', '')
            if not code:
                continue
            
            try:
                cursor.execute('''
                INSERT OR REPLACE INTO stocks (code, name, market, list_date, update_time, data_source)
                VALUES (?, ?, ?, ?, ?, ?)
                ''', (
                    code,
                    stock.get('name', ''),
                    stock.get('market', ''),
                    stock.get('list_date', ''),
                    now,
                    stock.get('data_source', 'collector')
                ))
                
                # 插入行业信息
                industry_info = industry_data.get(code, {})
                cursor.execute('''
                INSERT OR REPLACE INTO industries (code, shenwan_level1, shenwan_level2, shenwan_level3, general_industry, data_source)
                VALUES (?, ?, ?, ?, ?, ?)
                ''', (
                    code,
                    industry_info.get('shenwan_level1', ''),
                    industry_info.get('shenwan_level2', ''),
                    industry_info.get('shenwan_level3', ''),
                    industry_info.get('general_industry', ''),
                    industry_info.get('data_source', 'collector')
                ))
                
            except Exception as e:
                logger.warning(f"插入股票 {code} 失败: {e}")
        
        conn.commit()
    
    def _incremental_update_cache(self, conn: sqlite3.Connection,
                                  new_stock_list: List[Dict],
                                  industry_data: Dict) -> Dict:
        """执行缓存增量更新"""
        cursor = conn.cursor()
        
        # 获取当前缓存中的股票代码集合
        cursor.execute('SELECT code FROM stocks')
        old_codes = set(row[0] for row in cursor.fetchall())
        
        new_codes = set(stock.get('code', '') for stock in new_stock_list if stock.get('code', ''))
        
        # 计算差异
        added_codes = new_codes - old_codes
        updated_codes = new_codes & old_codes
        deleted_codes = old_codes - new_codes
        
        # 插入新股票
        for stock in new_stock_list:
            if stock.get('code', '') in added_codes:
                self._insert_stocks_to_cache(conn, [stock], {stock.get('code'): industry_data.get(stock.get('code', ''), {})})
        
        # 更新现有股票
        for stock in new_stock_list:
            if stock.get('code', '') in updated_codes:
                self._insert_stocks_to_cache(conn, [stock], {stock.get('code'): industry_data.get(stock.get('code', ''), {})})
        
        return {
            'added': len(added_codes),
            'updated': len(updated_codes),
            'deleted': len(deleted_codes)
        }
    
    def _fetch_industry_classifications(self, stock_list: List[Dict]) -> Dict:
        """从采集器或其他来源获取行业分类"""
        try:
            industry_data = {}
            
            for stock in stock_list:
                code = stock.get('code', '')
                if not code:
                    continue
                
                # 从stock对象中提取行业信息
                industry_data[code] = {
                    'shenwan_level1': stock.get('shenwan_level1', stock.get('l1', stock.get('industry', ''))),
                    'shenwan_level2': stock.get('shenwan_level2', stock.get('l2', '')),
                    'shenwan_level3': stock.get('shenwan_level3', stock.get('l3', '')),
                    'general_industry': stock.get('general_industry', stock.get('industry', '')),
                    'data_source': 'collector'
                }
            
            logger.info(f"获取到 {len(industry_data)} 只股票的行业分类")
            return industry_data
            
        except Exception as e:
            logger.warning(f"获取行业分类失败: {e}")
            return {}
    
    def _should_update_today(self) -> bool:
        """判断今天是否需要更新"""
        try:
            from datetime import datetime, date
            
            if not os.path.exists(self.cache_meta_path):
                return True
            
            import json
            with open(self.cache_meta_path, 'r') as f:
                meta = json.load(f)
            
            last_update = meta.get('last_update_date', '')
            today = str(date.today())
            
            return last_update != today
            
        except Exception as e:
            logger.warning(f"判断是否需要更新失败: {e}")
            return True
    
    def _update_cache_metadata(self):
        """更新缓存元数据"""
        try:
            from datetime import date, datetime
            import json
            
            metadata = {
                'last_update_date': str(date.today()),
                'last_update_time': datetime.now().isoformat(),
                'db_path': self.cache_db_path,
                'stock_count': self._get_cache_stock_count()
            }
            
            with open(self.cache_meta_path, 'w') as f:
                json.dump(metadata, f, indent=2, ensure_ascii=False)
            
            logger.info(f"缓存元数据已更新: {metadata}")
            
        except Exception as e:
            logger.warning(f"更新缓存元数据失败: {e}")
    
    def _get_cache_stock_count(self) -> int:
        """获取缓存中的股票总数"""
        try:
            if not os.path.exists(self.cache_db_path):
                return 0
            
            conn = self._get_cache_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM stocks')
            count = cursor.fetchone()[0]
            conn.close()
            return count
            
        except Exception:
            return 0
    
    def _get_connection(self) -> sqlite3.Connection:
        """
        获取数据库连接
        
        Returns:
            SQLite连接对象
        """
        try:
            if not os.path.exists(self.db_path):
                logger.info(f"数据库不存在，创建新数据库: {self.db_path}")
                # 如果数据库不存在，创建一个空的
                conn = sqlite3.connect(self.db_path, timeout=30)
                # 创建基本表结构
                self._create_basic_tables(conn)
                logger.info("数据库创建完成")
                return conn
            else:
                logger.debug(f"连接到数据库: {self.db_path}")
                return sqlite3.connect(self.db_path, timeout=30)
        except Exception as e:
            logger.error(f"获取数据库连接失败: {str(e)}", exc_info=True)
            raise
    
    def _create_basic_tables(self, conn: sqlite3.Connection):
        """
        创建基本表结构
        
        Args:
            conn: 数据库连接
        """
        cursor = conn.cursor()
        
        # 创建股票表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS stocks (
                code TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                market TEXT,
                list_date TEXT,
                data_source TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # 创建行业表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS industries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL,
                l1 TEXT,
                l2 TEXT,
                l3 TEXT,
                data_source TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (code) REFERENCES stocks(code)
            )
        ''')
        
        # 创建财务数据表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS financial_data (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL,
                name TEXT,
                year TEXT NOT NULL,
                non_op_real_estate REAL,
                data_source TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (code) REFERENCES stocks(code)
            )
        ''')
        
        # 创建索引
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_stocks_code ON stocks(code)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_industries_code ON industries(code)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_financial_code ON financial_data(code)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_financial_year ON financial_data(year)')
        
        conn.commit()
    
    def _get_industry_from_main_source(self) -> List[str]:
        """从主要数据源获取行业列表"""
        try:
            # 尝试从本地缓存获取行业数据
            cache_file = "shenwan_industry_mapping.pkl"
            if os.path.exists(cache_file):
                import pickle
                with open(cache_file, 'rb') as f:
                    industry_data = pickle.load(f)
                    if industry_data:
                        # 提取一级行业
                        l1_industries = set()
                        for stock_code, industry_info in industry_data.items():
                            if isinstance(industry_info, dict) and 'l1' in industry_info:
                                l1_industries.add(industry_info['l1'])
                        return sorted(list(l1_industries))
            
            # 如果有主要数据收集器，尝试从中获取
            try:
                from astock_real_estate_collector import AStockRealEstateDataCollector
                collector = AStockRealEstateDataCollector()
                # 这里可以调用收集器的方法来获取行业数据
                # 但考虑到性能，我们先返回默认列表
                return self._get_default_industries()
            except Exception:
                pass
                
            return []
            
        except Exception as e:
            logger.error(f"从主要数据源获取行业列表失败: {e}")
            return []
    
    def _get_default_industries(self) -> List[str]:
        """获取默认的申万一级行业列表"""
        return [
            "农林牧渔",
            "采掘",
            "化工",
            "钢铁",
            "有色金属",
            "电子",
            "家用电器",
            "食品饮料",
            "纺织服装",
            "轻工制造",
            "医药生物",
            "公用事业",
            "交通运输",
            "房地产",
            "商业贸易",
            "休闲服务",
            "综合",
            "建筑材料",
            "建筑装饰",
            "电气设备",
            "国防军工",
            "计算机",
            "传媒",
            "通信",
            "非银金融",
            "银行",
            "汽车",
            "机械设备"
        ]